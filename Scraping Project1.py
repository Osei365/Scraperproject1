#!/usr/bin/env python
# coding: utf-8

# In[2]:


import urllib3
import time
import requests
import xlsxwriter
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from selenium.common.exceptions import NoSuchElementException



def scraper(filename, sheetname):
    
    # Initiate selenium webdriver
    driver = webdriver.Chrome('chromedriver')
    #Getting the links to all the pages on the website
    
    all_page_links = []
    for n in range(25):
        if n == 0:
            all_page_links.append('https://www.naatp.org/resources/addiction-industry-directory')
        else:
            all_page_links.append('https://www.naatp.org/resources/addiction-industry-directory?page={}'.format(str(n)))
    
        

     
    #getting specific classes for row headers
    
    http = urllib3.PoolManager()
    
    # sending a get request to this website
    resp = http.request('GET', 'https://www.naatp.org/resources/addiction-industry-directory/3832/1-method-center')
    #page = requests.get('https://www.naatp.org/resources/addiction-industry-directory/3832/1-method-center')
    
    soup = BeautifulSoup(resp.data, 'html.parser')

    div1 = soup.find('div', class_ = 'views-row views-row-1 views-row-odd views-row-first views-row-last')
    div2 = div1.find_all('div')

    all_classes=[]
    for div in div2[2:-4]:
        all_classes.append(' '.join(div.get('class')))
    




    # Deriving Excel headers fron the classes scraped from websites
    headers = []
    for strings in all_classes:
        strings.split()
        headers.append(''.join(strings.split()[1])[12:])
    all_headers = ['Company_name', 'Accredited'] + headers
    



    # Creating an Excel Workbook to store data
    Workbook = xlsxwriter.Workbook(filename)
    sheet = Workbook.add_worksheet(sheetname)

    # Adding Bold format to the headers
    bold = Workbook.add_format({'bold': True})

    # All required Column  values according to Excel structure
    columns = 'A B C D E F G H I J K L M N O P Q R S T U V'.split()

    # Inserting the headers in the first row of each column
    for i, header in zip(columns, all_headers):
        sheet.write('{}1'.format(i), header, bold)

    Workbook.close()
   


   
    
    
    # Since Xlsxwriter just writes, we use Openpyxl to insert values into cells in the created Excel Workbook
    xfile= openpyxl.load_workbook(filename)
    sheet = xfile.get_sheet_by_name(sheetname)

    # The id of the element containing the field content is 'field-content'.
    # We define it for all 20 fields
    field_contents = ['field-content'] * 20

  
    # Iterating through all the pages
    for page_link in all_page_links:

        # sending a request to the page link
        driver.get(page_link)

        # getting the URLS to the fields on the page that will serve as Excel rows
        xpath = "//td[@class='views-field views-field-display-name']"
        all_links= driver.find_elements_by_xpath(xpath)
        all_links_list= [link.find_element_by_tag_name('a').get_attribute('href') for link in all_links]

        for link in all_links_list:
            n = sheet.max_row
            print(n)
            driver.get(link)
            try:
                sheet['A{}'.format(str(n+1))]= driver.find_element_by_tag_name('h1').text
            except NoSuchElementException:
                sheet['A{}'.format(str(n+1))] = 'Not indicated'

            try:
                sheet['B{}'.format(str(n+1))] = driver.find_element_by_class_name('accredited-Yes').text
            except NoSuchElementException:
                sheet['B{}'.format(str(n+1))] = ' Not Accredited'

            

            for sec, col in zip(all_classes, columns[2:]):
                try:
                    #header = driver.find_element_by_xpath('''//div[@class={}]//span[@class={}]".format(sec, pri)''').text
                    field = driver.find_element_by_xpath('''//div[@class='{}']//span[@class='field-content']'''.format(sec)).text
                    sheet['{}{}'.format(col, str(n+1))] = field
                    print(field)
                except NoSuchElementException:
                    sheet['{}{}'.format(col, str(n+1))] = 'Not Indicated'

            time.sleep(4)
        time.sleep(2)

    xfile.save('practice.xlsx')   
    
if __name__ == '__main__':
    
    scraper('practice.xlsx', 'Sheet1')


# In[ ]:




